import re
import unicodedata
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ------------------------------------------------------------
# Grundeinstellungen
# ------------------------------------------------------------
SHEET_NAME = "Druck Fahrer"

DAY_DEFINITIONS = [
    ("Sonntag", 4, 5),
    ("Montag", 6, 7),
    ("Dienstag", 8, 9),
    ("Mittwoch", 10, 11),
    ("Donnerstag", 12, 13),
    ("Freitag", 14, 15),
    ("Samstag", 16, 17),
]

SECTION_ORDER = [
    "Fahrer",
    "Hoffahrer + Waschteam",
    "Ausbildung zum Berufskraftfahrer",
    "Aushilfsfahrer",
]

SECTION_MARKERS = [
    ("Hoffahrer + Waschteam", ["hoffahrer", "waschteam"]),
    ("Ausbildung zum Berufskraftfahrer", ["ausbildung", "berufskraftfahrer"]),
    ("Aushilfsfahrer", ["aushilfsfahrer"]),
]

# Diese Disponenten werden nicht mehr in die Meldung übernommen.
EXCLUDED_PEOPLE = {
    ("aniol", "przemyslaw"),
    ("carstensen", "martin"),
    ("lau", "eike"),
    ("ohlenroth", "nadja"),
    ("packmohr", "gina"),
    ("pham manh", "chris"),
    ("schulz", "julian"),
}

# Diese Speditions-/Dienstleister-Zeilen werden ebenfalls ignoriert.
# Wichtig: "Maas Michael" als Fahrer bleibt drin. Ausgeschlossen wird nur "Sped. Maas".
EXCLUDED_LABELS = {
    "spedition meyer 1",
    "spedition meyer 2 36er",
    "spedition meyer 3",
    "spedition meyer 4",
    "spedition meyer 5",
    "spedition meyer sz",
    "paasch reinke 1",
    "paasch reinke 2",
    "paasch reinke 3",
    "devries 1",
    "de vries 1",
    "spedition ihde",
    "insellogistik 1",
    "insellogistik 2",
    "zippel logistik t1",
    "zippel logistik t2",
    "zippel logistik t3",
    "ch holtz t1",
    "ch holtz t2",
    "ch holtz t3",
    "kudex 1",
    "kudex 2",
    "kudex 3",
    "kudex 4",
    "pfenning 1",
    "pfenning 2",
    "t d",
    "sped maas",
    "nordfrost",
    "emons",
    "thermotraffic",
}

# Diese Begriffe werden übernommen. Abwandlungen werden mit erkannt,
# zum Beispiel krank, krankmeldung, urlaub, urlaubstag, modulschulung usw.
RELEVANT_PATTERNS = [
    r"\bonboarding\w*\b",
    r"\burlaub\w*\b",
    r"\bsonder\s*urlaub\w*\b",
    r"\bsonderurlaub\w*\b",
    r"\bkrank\w*\b",
    r"\bausgleich\w*\b",
    r"\belternzeit\w*\b",
    r"\bberufsschule\w*\b",
    r"\bfahrschule\w*\b",
    r"\bhome\s*[-]?\s*office\b",
    r"\bhomeoffice\b",
    r"\b\w*schulung\w*\b",
    r"\bdienst\s*reise\w*\b",
    r"\bdienstreise\w*\b",
    r"\bseminar\w*\b",
    r"\bfortbildung\w*\b",
    r"\bkur\w*\b",
    r"\breha\w*\b",
    r"\bmeeting\s*[-/]?\s*dispo\b",
    r"\bmodul\w*\b",
    r"\bfreigestellt\w*\b",
]

# Reine Uhrzeiten und reine Tournummern werden nicht übernommen.
TIME_PATTERN = re.compile(r"^\d{1,2}\s*[:.]\s*\d{2}$")
PLAIN_NUMBER_PATTERN = re.compile(r"^\d+(?:[,.]\d+)?$")


# ------------------------------------------------------------
# Hilfsfunktionen für Text und Erkennung
# ------------------------------------------------------------
def normalize_text(value) -> str:
    """Text robust vergleichbar machen: klein, ohne Akzente, saubere Leerzeichen."""
    if value is None:
        return ""

    text = str(value).strip()
    if text.lower() in {"nan", "none", "nat"}:
        return ""

    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.lower()
    text = re.sub(r"\s+", " ", text).strip()
    return text


def key_text(value) -> str:
    """Textschlüssel ohne Satzzeichen für Ausschlusslisten."""
    text = normalize_text(value)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def clean_output(value) -> str:
    """Originalwert für die Ausgabe säubern."""
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "nat"}:
        return ""
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def safe_cell(df: pd.DataFrame, row_idx: int, col_idx: int):
    if row_idx < 0 or col_idx < 0:
        return None
    if row_idx >= len(df.index) or col_idx >= len(df.columns):
        return None
    return df.iat[row_idx, col_idx]


def is_time_value(value) -> bool:
    text = normalize_text(value)
    if not text:
        return False
    return bool(TIME_PATTERN.fullmatch(text))


def is_plain_number(value) -> bool:
    text = normalize_text(value)
    if not text:
        return False
    return bool(PLAIN_NUMBER_PATTERN.fullmatch(text))


def is_relevant_entry(value) -> bool:
    text = normalize_text(value)

    if not text:
        return False

    # Überschriften, Platzhalter, Uhrzeiten und Tournummern ignorieren.
    if text in {"tour", "uhrzeit", "name", "vorname", "n.a", "n. a.", "na", "n/a"}:
        return False

    if is_time_value(text) or is_plain_number(text):
        return False

    return any(re.search(pattern, text) for pattern in RELEVANT_PATTERNS)


def row_has_excluded_label(row_values) -> bool:
    row_key = key_text(" ".join(clean_output(value) for value in row_values if clean_output(value)))
    if not row_key:
        return False

    return any(label in row_key for label in EXCLUDED_LABELS)


def get_section_marker(row_values) -> str | None:
    row_text = " ".join(normalize_text(value) for value in row_values if normalize_text(value))
    if not row_text:
        return None

    for section_name, marker_words in SECTION_MARKERS:
        if all(word in row_text for word in marker_words):
            return section_name

    return None


def is_employee_row(df: pd.DataFrame, row_idx: int) -> bool:
    row_values = df.iloc[row_idx].tolist()

    if row_has_excluded_label(row_values):
        return False

    lastname = normalize_text(safe_cell(df, row_idx, 1))
    firstname = normalize_text(safe_cell(df, row_idx, 2))
    personal_number = normalize_text(safe_cell(df, row_idx, 3))

    if (lastname, firstname) in EXCLUDED_PEOPLE:
        return False

    if not lastname or not firstname:
        return False

    if lastname in {"name", "datum", "zusatzliche arbeitszeit", "unterschrift mitarbeiter", "unterschrift fuhrparkleiter"}:
        return False

    if firstname in {"vorname", "tour"}:
        return False

    # Leere Platzhalter in der Ausbildung haben teilweise nur Nummern ohne Namen.
    if lastname.isdigit() or firstname.isdigit():
        return False

    # In Spalte D steht normalerweise Personalnummer oder eine andere Zuordnung.
    # Wenn die Spalte komplett leer ist, ist es meistens keine echte Fahrerzeile.
    if not personal_number:
        return False

    return True


def is_tour_row_for_employee(df: pd.DataFrame, row_idx: int) -> bool:
    return normalize_text(safe_cell(df, row_idx, 3)) == "tour"


def extract_day_entries(df: pd.DataFrame, driver_row_idx: int, col1: int, col2: int) -> str:
    """Für einen Wochentag relevante Texte aus Fahrerzeile und Tourzeile lesen."""
    rows_to_check = [driver_row_idx]
    if is_tour_row_for_employee(df, driver_row_idx + 1):
        rows_to_check.append(driver_row_idx + 1)

    entries: list[str] = []
    seen: set[str] = set()

    for row_idx in rows_to_check:
        for col_idx in (col1, col2):
            value = safe_cell(df, row_idx, col_idx)
            if not is_relevant_entry(value):
                continue

            output = clean_output(value)
            key = normalize_text(output)
            if output and key not in seen:
                entries.append(output)
                seen.add(key)

    return " / ".join(entries)


# ------------------------------------------------------------
# Daten aus Druck Fahrer extrahieren
# ------------------------------------------------------------
def extract_grouped_work_data(df: pd.DataFrame) -> pd.DataFrame:
    result_rows = []
    current_section = "Fahrer"

    for row_idx in range(len(df.index)):
        marker = get_section_marker(df.iloc[row_idx].tolist())
        if marker:
            current_section = marker
            continue

        if not is_employee_row(df, row_idx):
            continue

        lastname = clean_output(safe_cell(df, row_idx, 1)).title()
        firstname = clean_output(safe_cell(df, row_idx, 2)).title()

        row = {
            "Kategorie": current_section,
            "Nachname": lastname,
            "Vorname": firstname,
        }

        has_relevant_entry = False
        for weekday, col1, col2 in DAY_DEFINITIONS:
            entry = extract_day_entries(df, row_idx, col1, col2)
            row[weekday] = entry
            if entry:
                has_relevant_entry = True

        # Nur Fahrer übernehmen, bei denen mindestens ein relevanter Eintrag gefunden wurde.
        if has_relevant_entry:
            result_rows.append(row)

    extracted = pd.DataFrame(result_rows)

    if extracted.empty:
        columns = ["Kategorie", "Nachname", "Vorname"] + [weekday for weekday, _, _ in DAY_DEFINITIONS]
        return pd.DataFrame(columns=columns)

    # Reihenfolge festlegen, damit die Blöcke sauber untereinander stehen.
    extracted["_section_order"] = extracted["Kategorie"].apply(
        lambda value: SECTION_ORDER.index(value) if value in SECTION_ORDER else 999
    )
    extracted = extracted.sort_values(
        by=["_section_order", "Nachname", "Vorname"],
        kind="stable",
    ).drop(columns=["_section_order"])

    return extracted.reset_index(drop=True)


# ------------------------------------------------------------
# Datumszeile lesen
# ------------------------------------------------------------
def create_header_with_dates(df: pd.DataFrame) -> list[str]:
    dates: list[str] = []

    for _, col1, _ in DAY_DEFINITIONS:
        value = safe_cell(df, 1, col1)
        parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")

        if pd.isna(parsed):
            dates.append("")
        else:
            dates.append(parsed.strftime("%d.%m.%Y"))

    return dates


def get_report_week(dates: list[str]) -> int:
    first_date = pd.to_datetime(dates[0], format="%d.%m.%Y", errors="coerce")
    if pd.isna(first_date):
        raise ValueError("Das erste Datum in der Datumszeile konnte nicht gelesen werden.")

    # Die Druck-Fahrer-Datei beginnt am Sonntag. Für die Kalenderwoche gilt die Folgewoche ab Montag.
    return int(first_date.isocalendar().week) + 1


def apply_dates_to_columns(df: pd.DataFrame, dates: list[str]) -> pd.DataFrame:
    output = df.copy()
    rename_map = {}

    for (weekday, _, _), date_text in zip(DAY_DEFINITIONS, dates):
        rename_map[weekday] = f"{weekday} ({date_text})" if date_text else weekday

    return output.rename(columns=rename_map)


# ------------------------------------------------------------
# Excel-Formatierung
# ------------------------------------------------------------
def style_excel(ws, report_week: int):
    title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    data_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    data_fill_light = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    aushilfe_fill_dark = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    aushilfe_fill_light = PatternFill(start_color="82E0AA", end_color="82E0AA", fill_type="solid")
    category_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    medium_border = Border(
        left=Side(style="medium", color="1F4E78"),
        right=Side(style="medium", color="1F4E78"),
        top=Side(style="medium", color="1F4E78"),
        bottom=Side(style="medium", color="1F4E78"),
    )

    # Titelzeile
    ws["A1"].value = f"Kalenderwoche: {report_week}"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = title_fill
    ws["A1"].border = medium_border
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    ws.row_dimensions[1].height = 30

    # Abteilung
    ws["A2"].value = "Abteilung: Fuhrpark - NFC"
    ws["A2"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].fill = title_fill
    ws["A2"].border = medium_border
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ws.max_column)
    ws.row_dimensions[2].height = 26

    # Tabellenkopf
    for row in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = header_fill
            cell.border = medium_border
    ws.row_dimensions[3].height = 32

    green_names = {
        ("kleiber", "lutz"),
        ("dammasch", "bernd"),
        ("linke", "erich"),
        ("steckel", "wolfgang"),
    }

    # Datenzeilen
    for row_idx in range(4, ws.max_row + 1):
        category = normalize_text(ws.cell(row=row_idx, column=1).value)
        lastname = normalize_text(ws.cell(row=row_idx, column=2).value)
        firstname = normalize_text(ws.cell(row=row_idx, column=3).value)

        base_fill = data_fill_light if row_idx % 2 == 0 else data_fill_white
        base_font = Font(size=10, color="2C3E50")

        if category == normalize_text("Aushilfsfahrer") or (lastname, firstname) in green_names:
            base_fill = aushilfe_fill_dark if row_idx % 2 == 0 else aushilfe_fill_light
            base_font = Font(size=10, color="FFFFFF" if row_idx % 2 == 0 else "2C3E50", bold=True)

        max_text_length = 0
        for cell in ws[row_idx]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = base_font
            cell.fill = base_fill
            cell.border = thin_border
            if cell.value:
                max_text_length = max(max_text_length, len(str(cell.value)))

        # Kategorie optisch hervorheben
        ws.cell(row=row_idx, column=1).fill = category_fill
        ws.cell(row=row_idx, column=1).font = Font(size=10, color="1F4E78", bold=True)

        # Zeilenhöhe an Inhalt anpassen, aber nicht übertreiben.
        if max_text_length > 55:
            ws.row_dimensions[row_idx].height = 44
        elif max_text_length > 35:
            ws.row_dimensions[row_idx].height = 34
        else:
            ws.row_dimensions[row_idx].height = 24

    # Spaltenbreite sauber setzen
    column_min_widths = {
        1: 28,  # Kategorie
        2: 20,  # Nachname
        3: 18,  # Vorname
    }

    for col_idx, col in enumerate(ws.columns, start=1):
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        if col_idx <= 3:
            min_width = column_min_widths[col_idx]
            max_width = 34
        else:
            min_width = 22
            max_width = 42

        calculated_width = max_length + 3
        adjusted_width = min(max(calculated_width, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted_width

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = ws.dimensions


# ------------------------------------------------------------
# Streamlit App
# ------------------------------------------------------------
st.set_page_config(page_title="Wochenarbeitsbericht Fuhrpark", layout="wide")
st.title("Wochenarbeitsbericht Fuhrpark")
st.info("Disponenten und Speditions-/Dienstleister-Zeilen werden nicht übernommen. Ausgewertet werden relevante Einträge aus dem Blatt 'Druck Fahrer'.")

uploaded_file = st.file_uploader("Lade eine Excel-Datei hoch", type=["xlsx"])

if uploaded_file:
    progress_bar = st.progress(0)
    progress_status = st.empty()

    progress_status.text("Lade Excel-Datei...")
    try:
        wb = load_workbook(uploaded_file, data_only=True)
    except Exception as exc:
        st.error(f"Die Excel-Datei konnte nicht gelesen werden: {exc}")
        st.stop()

    if SHEET_NAME not in wb.sheetnames:
        st.error(f"Das Tabellenblatt '{SHEET_NAME}' wurde nicht gefunden.")
        st.stop()

    sheet = wb[SHEET_NAME]
    data = pd.DataFrame(sheet.values)
    progress_status.text("Excel-Daten geladen.")
    progress_bar.progress(30)

    progress_status.text("Lese Fahrer, Hoffahrer, Ausbildung und Aushilfsfahrer...")
    extracted_data = extract_grouped_work_data(data)
    progress_bar.progress(60)

    if extracted_data.empty:
        st.warning("Es wurden keine relevanten Einträge gefunden.")
        st.stop()

    dates = create_header_with_dates(data)

    try:
        report_week = get_report_week(dates)
    except ValueError as exc:
        st.error(str(exc))
        st.stop()

    output_data = apply_dates_to_columns(extracted_data, dates)

    excel_filename = f"Fuhrpark_Meldung_KW_{report_week}.xlsx"

    progress_status.text("Erstelle Excel-Datei...")
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        output_data.to_excel(writer, index=False, sheet_name="Wochenübersicht", startrow=2)
        ws = writer.sheets["Wochenübersicht"]
        style_excel(ws, report_week)

    excel_data = output.getvalue()

    progress_status.text("Fertig!")
    progress_bar.progress(100)
    st.success("Verarbeitung abgeschlossen.")

    st.subheader("Vorschau")
    st.dataframe(output_data, use_container_width=True, hide_index=True)

    st.download_button(
        label="Download als Excel",
        data=excel_data,
        file_name=excel_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
